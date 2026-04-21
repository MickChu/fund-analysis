# -*- coding: utf-8 -*-
"""
import_alipay_csv.py
从支付宝9个CSV文件(GBK编码)解析基金买卖记录，输出买入+卖出两个Excel。
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

from collections import Counter
from datetime import datetime
from pathlib import Path
from logger import get_logger

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR    = PROJECT_DIR / "数据"
OUTPUT_DIR  = PROJECT_DIR / "输出"

# 初始化日志
log = get_logger('import_alipay_csv')

FUND_MAP = {
    "000051": "华夏沪深300ETF联接A",
    "001717": "工银前沿医疗股票A",
    "006479": "广发纳指100ETF联接C",
    "007760": "景顺沪港深低波指数C",
    "008702": "华夏黄金ETF联接C",
    "010339": "国投瑞银新能源混合C",
    "012323": "华宝医疗ETF联接C",
    "014162": "万家人工智能混合C",
    "020640": "广发半导体ETF联接C",
    "022460": "易方达A500ETF联接C",
    "023918": "华夏自由现金流ETF联接C",
    "024617": "大成自由现金流ETF联接C",
    "162216": "宏利中证500增强",
}

# 用unicode代码点定义所有中文字符，避免编码问题
_HUA = chr(21326)    # 华
_XIA = chr(22799)    # 夏
_HU = chr(27818)     # 沪
_SHEN = chr(28145)   # 深
_LIAN = chr(32852)   # 联
_JIE = chr(25509)    # 接
_TIAN = chr(22825)   # 天
_HONG = chr(24344)   # 弘
_GONG = chr(24037)   # 工
_YIN = chr(38134)    # 银
_RUI = chr(29790)    # 瑞
_XIN = chr(20449)    # 信
_QIAN = chr(21069)   # 前
_YAN = chr(27839)    # 沿
_YI = chr(21307)     # 医
_LIAO = chr(30103)   # 疗
_GU = chr(32929)     # 股
_PIAO = chr(31080)   # 票
_GUANG = chr(24191)  # 广
_FA = chr(21457)     # 发
_NA = chr(32435)     # 纳
_SI = chr(26031)     # 斯
_DA = chr(36798)     # 达
_KE = chr(20811)     # 克
_JING = chr(26223)   # 景
_SHUN = chr(39034)   # 顺
_CHANG = chr(38271)  # 长
_CHENG = chr(22478)  # 城
_ZHONG = chr(20013)  # 中
_ZHENG = chr(35777)  # 证
_GANG = chr(28207)   # 港
_HONG2 = chr(32418)  # 红
_LI = chr(21033)     # 利
_CHENG2 = chr(25104) # 成
_DI = chr(20302)     # 低
_BO = chr(27874)     # 波
_DONG = chr(21160)   # 动
_HUANG = chr(40644)  # 黄
_JIN = chr(37329)    # 金
_GUO = chr(22269)    # 国
_TOU = chr(25237)    # 投
_XIN2 = chr(26032)   # 新
_NENG = chr(33021)   # 能
_YUAN = chr(28304)   # 源
_HUN = chr(28151)    # 混
_HE = chr(21512)     # 合
_BAO = chr(23453)    # 宝
_WAN = chr(19975)    # 万
_JIA = chr(23478)    # 家
_REN = chr(20154)    # 人
_ZHI = chr(26234)    # 智
_BAN = chr(21322)    # 半
_DAO = chr(23548)    # 导
_TI = chr(20307)     # 体
_YI2 = chr(26131)    # 易
_FANG = chr(26041)   # 方
_ZI = chr(33258)     # 自
_YOU = chr(30001)    # 由
_XIAN = chr(29616)   # 现
_DA2 = chr(22823)    # 大
_HONG3 = chr(23439)  # 宏
_ZENG = chr(22686)   # 增
_QIANG = chr(24378)  # 强
_ZHI2 = chr(25351)   # 指
_SHU = chr(25968)    # 数
_LIU = chr(27969)    # 流

# 关键词字典 (用chr()拼接中文)
KW2CODE = {
    # 000051 沪深300
    _HUA + _XIA + _HU + _SHEN + "300ETF" + _LIAN + _JIE: "000051",
    _TIAN + _HONG + _HU + _SHEN + "300ETF" + _LIAN + _JIE: "000051",
    _HU + _SHEN + "300ETF" + _LIAN + _JIE: "000051",
    
    # 001717 前沿医疗
    _GONG + _YIN + _RUI + _XIN + _QIAN + _YAN + _YI + _LIAO + _GU + _PIAO: "001717",
    _GONG + _YIN + _QIAN + _YAN + _YI + _LIAO: "001717",
    _QIAN + _YAN + _YI + _LIAO + _GU + _PIAO: "001717",
    
    # 006479 纳指100
    _GUANG + _FA + _NA + _SI + _DA + _KE + "100" + _ZHI2 + _SHU: "006479",
    _GUANG + _FA + _NA + _SI + _DA + _KE + "100ETF" + _LIAN + _JIE: "006479",
    _NA + _SI + _DA + _KE + "100" + _ZHI2 + _SHU: "006479",
    _NA + _SI + _DA + _KE + "100ETF": "006479",
    _NA + _ZHI2 + "100": "006479",
    
    # 007760 沪港深低波
    _JING + _SHUN + _CHANG + _CHENG + _ZHONG + _ZHENG + _HU + _GANG + _SHEN + _HONG2 + _LI + _CHENG2 + _CHANG + _DI + _BO + _DONG: "007760",
    _HU + _GANG + _SHEN + _HONG2 + _LI + _CHENG2 + _CHANG + _DI + _BO + _DONG: "007760",
    _HU + _GANG + _SHEN + _DI + _BO + _DONG: "007760",
    
    # 008702 黄金ETF
    _HUA + _XIA + _HUANG + _JIN + "ETF" + _LIAN + _JIE: "008702",
    _HUANG + _JIN + "ETF" + _LIAN + _JIE + "C": "008702",
    
    # 010339 新能源
    _GUO + _TOU + _RUI + _YIN + _XIN2 + _NENG + _YUAN + _HUN + _HE: "010339",
    _XIN2 + _NENG + _YUAN + _HUN + _HE + "C": "010339",
    
    # 012323 医疗ETF
    _HUA + _BAO + _ZHONG + _ZHENG + _YI + _LIAO + "ETF" + _LIAN + _JIE: "012323",
    _YI + _LIAO + "ETF" + _LIAN + _JIE + "C": "012323",
    
    # 014162 人工智能
    _WAN + _JIA + _REN + _GONG + _ZHI + _NENG + _HUN + _HE: "014162",
    _REN + _GONG + _ZHI + _NENG + _HUN + _HE + "C": "014162",
    
    # 020640 半导体
    _HUA + _XIA + _GUO + _ZHENG + _BAN + _DAO + _TI + "ETF" + _LIAN + _JIE: "020640",
    _BAN + _DAO + _TI + "ETF" + _LIAN + _JIE: "020640",
    
    # 022460 A500
    _YI2 + _FANG + _DA + _ZHONG + _ZHENG + "A500ETF" + _LIAN + _JIE: "022460",
    "A500ETF" + _LIAN + _JIE + "C": "022460",
    
    # 023918 自由现金流
    _HUA + _XIA + _GUO + _ZHENG + _ZI + _YOU + _XIAN + _JIN + _LIU + "ETF" + _LIAN + _JIE: "023918",
    _ZI + _YOU + _XIAN + _JIN + _LIU + "ETF" + _LIAN + _JIE + "C": "023918",
    
    # 024617 大成自由现金流
    _DA2 + _CHENG2 + _ZHONG + _ZHENG + _ZI + _YOU + _XIAN + _JIN + _LIU + "ETF" + _LIAN + _JIE: "024617",
    
    # 162216 中证500增强
    _HONG3 + _LI + _ZHONG + _ZHENG + "500" + _ZHI2 + _SHU + _ZENG + _QIANG: "162216",
    _ZHONG + _ZHENG + "500" + _ZHI2 + _SHU + _ZENG + _QIANG: "162216",
    _ZHONG + _ZHENG + "500" + _ZENG + _QIANG: "162216",
}

ALL_KWS = sorted(KW2CODE.keys(), key=len, reverse=True)

# 检查用unicode代码点
_MAYI = chr(34434) + chr(34433) + chr(36130) + chr(23500)  # 蚂蚁财富
_JIJIN = chr(22522) + chr(37329)  # 基金
_MAICHU = chr(21334) + chr(20986)  # 卖出
_SHUHUI = chr(36174) + chr(22238)  # 赎回
_YUEBAO = chr(20313) + chr(39069) + chr(23453)  # 余额宝
_HUOBI = chr(36135) + chr(24065)  # 货币
_MAIRU = chr(20080) + chr(20837)  # 买入

# 定义缺失的变量
_LIU = chr(27969)  # 流


def find_code(product: str):
    for kw in ALL_KWS:
        if kw in product:
            return KW2CODE[kw]
    return None


def parse_line(raw: str):
    line = raw.rstrip("\r\n")
    normalized = line.replace("\t", "")
    fields = [f.strip() for f in normalized.split(",")]
    while fields and fields[-1] == "":
        fields.pop()
    return fields


def parse_csv_files():
    csvs = sorted(DATA_DIR.glob("段*_*.csv"))
    if not csvs:
        print(f"错误: 在 {DATA_DIR} 中找不到CSV文件")
        return [], []

    buys, sells = [], []

    for csv_path in csvs:
        print(f"处理: {csv_path.name} ...", end="", flush=True)

        try:
            with open(csv_path, "r", encoding="gbk", errors="replace") as fh:
                lines = fh.readlines()
        except Exception as e:
            print(f"  读取失败: {e}")
            continue

        n_buy = n_sell = 0

        for raw in lines:
            if _MAYI not in raw and _JIJIN not in raw:
                continue

            f = parse_line(raw)
            if len(f) < 12:
                continue

            trade_date_raw = f[2]
            amount_raw     = f[9]
            direction      = f[10]
            status         = f[11]
            product        = f[8]

            if status != "交易成功":
                continue

            # 余额宝过滤：仅跳过非基金类余额宝交易，基金卖出到余额宝要保留
            # 余额宝过滤：仅跳过非基金类余额宝交易，基金卖出到余额宝要保留
            is_fund_sell = _MAICHU in product or _SHUHUI in product
            if (not is_fund_sell) and (_YUEBAO in product or _HUOBI in product):
                continue
            try:
                dt = datetime.strptime(trade_date_raw[:10], "%Y-%m-%d")
                date_str = dt.strftime("%Y-%m-%d")
            except (ValueError, TypeError):
                continue

            try:
                amount = float(amount_raw)
            except (ValueError, TypeError):
                continue

            code = find_code(product)
            if code not in FUND_MAP:
                continue

            name = FUND_MAP[code]

            if _MAICHU in product or _SHUHUI in product:
                t = "sell"
            elif _MAIRU in product or "-定投" in product:
                t = "buy"
            elif direction == "支出":
                t = "buy"
            elif direction == "收入":
                t = "sell"
            else:
                t = "buy"

            rec = {
                "date":    date_str,
                "code":    code,
                "name":    name,
                "amount":  round(amount, 2),
                "product": product,
                "source":  csv_path.name,
            }

            if t == "buy":
                buys.append(rec)
                n_buy += 1
            else:
                sells.append(rec)
                n_sell += 1

        print(f"  买入{n_buy}条 卖出{n_sell}条")

    buys.sort(key=lambda x: x["date"])
    sells.sort(key=lambda x: x["date"])
    return buys, sells


def write_xlsx(records, out_path, is_buy):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active

    if is_buy:
        ws.title = "买入记录"
        hdr = ["日期", "基金代码", "基金名称", "买入金额(元)", "申购费率(%)", "买入净值(可选)", "备注"]
        hdr_color = "4472C4"
    else:
        ws.title = "卖出记录"
        hdr = ["日期", "基金代码", "基金名称", "卖出金额(元)", "备注"]
        hdr_color = "C00000"

    ws.append(hdr)

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )
    hf    = PatternFill(start_color=hdr_color, end_color=hdr_color, fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for col in range(1, len(hdr) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = hf
        c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin

    if is_buy:
        for r in records:
            ws.append([r["date"], r["code"], r["name"], r["amount"], "", "", r["product"]])
        ws.column_dimensions["A"].width = 13
        ws.column_dimensions["B"].width = 11
        ws.column_dimensions["C"].width = 26
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 16
        ws.column_dimensions["G"].width = 45
    else:
        for r in records:
            ws.append([r["date"], r["code"], r["name"], r["amount"], r["product"]])
        ws.column_dimensions["A"].width = 13
        ws.column_dimensions["B"].width = 11
        ws.column_dimensions["C"].width = 26
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 45

    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("基金代码")
    ws2.append(["代码", "名称", "备注"])
    hf2 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col in [1, 2, 3]:
        c = ws2.cell(row=1, column=col)
        c.fill = hf2
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 26
    ws2.column_dimensions["C"].width = 40
    for code, name in FUND_MAP.items():
        ws2.append([code, name, ""])
    ws2.append([])
    ws2.append(["使用说明"])
    if is_buy:
        ws2.append(["1. 费率列(申购费率%)：C类填0，A类1折填0.12，留空默认0"])
        ws2.append(["2. 买入净值留空，程序运行时会自动从NAV缓存补全"])
        ws2.append(["3. 备注列为支付宝原始商品名称，可删除不相关的行"])
    else:
        ws2.append(["1. 卖出金额：实际赎回到账金额"])
        ws2.append(["2. 如有遗漏的卖出记录请手动补充"])

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"已保存: {out_path}  ({len(records)}条)")


def main():
    log.info("支付宝CSV导入器启动")
    print("=" * 60)
    print("支付宝CSV导入器")
    print("=" * 60)

    buys, sells = parse_csv_files()

    log.info(f"汇总: 买入{len(buys)}条 | 卖出{len(sells)}条")
    print(f"\n汇总: 买入{len(buys)}条 | 卖出{len(sells)}条")

    if buys:
        log.info("各基金买入统计:")
        print("\n各基金买入统计:")
        cnt = Counter(r["code"] for r in buys)
        for code in sorted(cnt):
            msg = f"  {code}  {FUND_MAP.get(code, '?'):30s}  {cnt[code]}条"
            log.info(msg)
            print(msg)

    if sells:
        log.info("各基金卖出统计:")
        print("\n各基金卖出统计:")
        cnt2 = Counter(r["code"] for r in sells)
        for code in sorted(cnt2):
            msg = f"  {code}  {FUND_MAP.get(code, '?'):30s}  {cnt2[code]}条"
            log.info(msg)
            print(msg)

    write_xlsx(buys,  OUTPUT_DIR / "支付宝_买入记录.xlsx", is_buy=True)
    write_xlsx(sells, OUTPUT_DIR / "支付宝_卖出记录.xlsx", is_buy=False)

    log.info("完成！请检查输出文件，补充费率列，并删除不相关基金行。")
    print("\n完成！请检查输出文件，补充费率列，并删除不相关基金行。")


if __name__ == "__main__":
    main()
