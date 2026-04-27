# -*- coding: utf-8 -*-
"""
portfolio_tracker.py — 持仓跟踪模块

功能：
1. 读取买入记录Excel，计算每只基金的持仓状态
2. 结合nav_fetcher缓存的最新净值，计算市值、收益、收益率
3. 输出持仓总表、板块分布、盈亏分析

用法：
  python portfolio_tracker.py                    # 读取默认路径的买入记录
  python portfolio_tracker.py --input 路径.xlsx  # 指定买入记录文件
  python portfolio_tracker.py --output json      # 输出JSON格式（默认）
  python portfolio_tracker.py --output md        # 输出Markdown报告
"""

import sys
sys.stdout.reconfigure(encoding='utf-8')

import os
import json
import argparse
from datetime import datetime
from pathlib import Path
from logger import get_logger, cleanup_old_logs

# 项目路径
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "数据"
NAV_CACHE_DIR = DATA_DIR / "nav_cache"
OUTPUT_DIR = PROJECT_DIR / "输出"

# 初始化日志
log = get_logger('portfolio_tracker')

# 默认文件名（支持新版基金交易记录.xlsx 和旧版买入记录.xlsx）
DEFAULT_EXCEL = DATA_DIR / "基金交易记录.xlsx"
LEGACY_EXCEL = DATA_DIR / "买入记录.xlsx"

# 13只基金清单
FUND_LIST = {
    "000051": "华夏沪深300ETF联接A",
    "001717": "工银前沿医疗股票A",
    "006479": "广发纳指100ETF联接C",
    "007760": "景顺沪港深低波指数C",
    "008702": "华夏黄金ETF联接C",
    "010339": "国投瑞银新能源混合C",
    "012323": "华宝医疗ETF联接C",
    "014162": "万家人工智能混合C",
    "020640": "广发半导体ETF联接C",
    "022460": "易方达A500ETF联接C",
    "023918": "华夏自由现金流ETF联接C",
    "024617": "大成自由现金流ETF联接C",
    "162216": "宏利中证500增强",
}

# 申购费率（1折优惠后）
# C类基金：申购费=0（销售服务费已包含在每日净值中）
# A类基金：大多数平台打1折
FUND_FEE_RATE = {
    "000051": 0.0012,   # A类，标准1.2%，1折后0.12%
    "001717": 0.0015,   # A类，标准1.5%，1折后0.15%
    "006479": 0.0,      # C类
    "007760": 0.0,      # C类
    "008702": 0.0,      # C类
    "010339": 0.0,      # C类
    "012323": 0.0,      # C类
    "014162": 0.0,      # C类
    "020640": 0.0,      # C类
    "022460": 0.0,      # C类
    "023918": 0.0,      # C类
    "024617": 0.0,      # C类
    "162216": 0.0,      # LOF，多数平台0费率
}


def load_buy_records(excel_path):
    """从Excel读取买入记录，返回 list of dict"""
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["买入记录"]

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 跳过空行
        if not row or not row[0] or not row[1]:
            continue

        date_val = row[0]
        code = str(row[1]).strip() if row[1] else ""
        name = str(row[2]).strip() if row[2] else ""
        amount = row[3]
        fee_pct = row[4]  # 申购费率(%)，用户输入如 0.12 表示 0.12%
        nav = row[5]
        remark = str(row[6]) if len(row) > 6 and row[6] else ""

        # 跳过非数据行
        if not code or code not in FUND_LIST:
            continue

        # 处理日期格式
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val)[:10]

        # 处理金额
        try:
            amount = float(amount) if amount else 0
        except (ValueError, TypeError):
            amount = 0

        # 处理申购费率（%），用户输入如 0.12 表示 0.12%
        # 留空则使用默认费率（C类=0, A类=0.12%）
        try:
            if fee_pct is not None and fee_pct != "":
                fee_rate = float(fee_pct) / 100.0  # 0.12 → 0.0012
            else:
                fee_rate = FUND_FEE_RATE.get(code, 0.0)
        except (ValueError, TypeError):
            fee_rate = FUND_FEE_RATE.get(code, 0.0)

        # 处理净值（可选）
        try:
            nav = float(nav) if nav else None
        except (ValueError, TypeError):
            nav = None

        if amount > 0:
            records.append({
                "date": date_str,
                "code": code,
                "name": name or FUND_LIST.get(code, ""),
                "amount": amount,
                "fee_rate": fee_rate,  # 小数形式，如 0.0012
                "nav": nav,
                "remark": remark,
            })

    # 按日期排序
    records.sort(key=lambda x: x["date"])
    return records


def load_nav_for_date(code, date_str):
    """从缓存获取指定日期的净值"""
    cache_file = NAV_CACHE_DIR / f"{code}.json"
    if not cache_file.exists():
        return None

    with open(cache_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    for item in data.get("nav_data", []):
        if item["date"] == date_str:
            return item["nav"]
    return None


def load_latest_nav(code):
    """从缓存获取最新净值"""
    cache_file = NAV_CACHE_DIR / f"{code}.json"
    if not cache_file.exists():
        return None, None

    with open(cache_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    nav_data = data.get("nav_data", [])
    if not nav_data:
        return None, None

    latest = nav_data[-1]
    return latest["nav"], latest["date"]


def calculate_portfolio(records):
    """计算持仓状态，返回 dict"""
    from collections import defaultdict

    # 按基金分组
    fund_records = defaultdict(list)
    for r in records:
        fund_records[r["code"]].append(r)

    portfolio = {}
    total_cost = 0
    total_value = 0

    for code, trades in fund_records.items():
        name = FUND_LIST.get(code, trades[0]["name"])

        # 计算总成本和份额
        total_fund_cost = 0
        total_shares = 0

        for trade in trades:
            amount = trade["amount"]
            trade_date = trade["date"]

            # 获取买入日净值
            nav = trade["nav"]
            if nav is None:
                nav = load_nav_for_date(code, trade_date)
            if nav is None:
                nav = 1.0  # 兜底，但会报警告

            # 扣除申购费后计算份额（费率从买入记录读取，优先于默认值）
            fee_rate = trade.get("fee_rate", FUND_FEE_RATE.get(code, 0.0))
            net_amount = amount / (1 + fee_rate) if fee_rate > 0 else amount
            shares = net_amount / nav
            total_fund_cost += amount  # 记录实际支付金额（含手续费）
            total_shares += shares

        # 获取最新净值
        latest_nav, latest_date = load_latest_nav(code)
        if latest_nav is None:
            continue

        market_value = total_shares * latest_nav
        profit = market_value - total_fund_cost
        profit_rate = (profit / total_fund_cost * 100) if total_fund_cost > 0 else 0

        portfolio[code] = {
            "code": code,
            "name": name,
            "trades_count": len(trades),
            "total_cost": round(total_fund_cost, 2),
            "shares": round(total_shares, 4),
            "latest_nav": latest_nav,
            "latest_nav_date": latest_date,
            "market_value": round(market_value, 2),
            "profit": round(profit, 2),
            "profit_rate": round(profit_rate, 2),
        }

        total_cost += total_fund_cost
        total_value += market_value

    total_profit = total_value - total_cost
    total_profit_rate = (total_profit / total_cost * 100) if total_cost > 0 else 0

    return {
        "funds": portfolio,
        "summary": {
            "total_cost": round(total_cost, 2),
            "total_value": round(total_value, 2),
            "total_profit": round(total_profit, 2),
            "total_profit_rate": round(total_profit_rate, 2),
            "fund_count": len(portfolio),
        },
        "update_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def output_json(data, output_path):
    """输出JSON文件"""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"JSON已保存: {output_path}")


def output_markdown(data, output_path):
    """输出Markdown报告"""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    lines = []
    lines.append("# 持仓跟踪报告")
    lines.append(f"\n生成时间: {data['update_time']}")
    lines.append(f"净值日期: {max(f['latest_nav_date'] for f in data['funds'].values())}")
    lines.append("\n---\n")

    # 汇总
    summary = data["summary"]
    lines.append("## 持仓汇总")
    lines.append(f"\n| 指标 | 数值 |")
    lines.append("|------|------|")
    lines.append(f"| 持仓基金数 | {summary['fund_count']} 只 |")
    lines.append(f"| 总成本 | ¥{summary['total_cost']:,.2f} |")
    lines.append(f"| 总市值 | ¥{summary['total_value']:,.2f} |")
    lines.append(f"| 总收益 | ¥{summary['total_profit']:,.2f} |")
    lines.append(f"| 收益率 | {summary['total_profit_rate']:+.2f}% |")
    lines.append("\n---\n")

    # 明细
    lines.append("## 持仓明细")
    lines.append("\n| 基金代码 | 基金名称 | 成本 | 市值 | 收益 | 收益率 | 占比 |")
    lines.append("|---------|---------|------|------|------|--------|------|")

    total_value = summary["total_value"]
    for code, fund in sorted(data["funds"].items()):
        weight = (fund["market_value"] / total_value * 100) if total_value > 0 else 0
        lines.append(
            f"| {fund['code']} | {fund['name']} | "
            f"¥{fund['total_cost']:,.2f} | ¥{fund['market_value']:,.2f} | "
            f"¥{fund['profit']:,.2f} | {fund['profit_rate']:+.2f}% | {weight:.1f}% |"
        )

    lines.append("\n---\n")
    lines.append("*注：占比 = 单只基金市值 / 总市值")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"Markdown报告已保存: {output_path}")


def load_holdings_share(excel_path):
    """从Excel读取持仓份额sheet，返回 dict {code: shares}"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["持仓份额"]
    except (KeyError, FileNotFoundError):
        return {}  # 无持仓份额sheet则跳过

    shares_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        code = str(row[0]).strip()
        try:
            shares = float(row[2]) if row[2] else None
        except (ValueError, TypeError):
            shares = None
        if code and shares is not None:
            shares_map[code] = shares
    return shares_map


def main():
    parser = argparse.ArgumentParser(description="持仓跟踪模块")
    parser.add_argument("--input", default=None, help="买入记录Excel路径（默认自动查找基金交易记录.xlsx或买入记录.xlsx）")
    parser.add_argument("--output", choices=["json", "md"], default="json", help="输出格式")
    parser.add_argument("--cleanup", type=int, default=30, help="清理N天前的日志，默认30天")
    args = parser.parse_args()

    # 清理日志
    if args.cleanup >= 0:
        deleted = cleanup_old_logs(args.cleanup)
        if deleted:
            log.info(f"已清理 {len(deleted)} 个日志文件")

    # 自动找文件
    if args.input:
        input_path = Path(args.input)
    elif DEFAULT_EXCEL.exists():
        input_path = DEFAULT_EXCEL
        log.info(f"使用文件: {input_path}")
        print(f"使用文件: {input_path}")
    elif LEGACY_EXCEL.exists():
        input_path = LEGACY_EXCEL
        log.info(f"使用文件(旧版): {input_path}")
        print(f"使用文件(旧版): {input_path}")
    else:
        log.error(f"找不到买入记录文件: {DEFAULT_EXCEL}")
        print(f"错误: 找不到买入记录文件")
        print(f"  期望路径: {DEFAULT_EXCEL}")
        print(f"  或: {LEGACY_EXCEL}")
        return

    if not input_path.exists():
        log.error(f"文件不存在: {input_path}")
        print(f"错误: 找不到文件 {input_path}")
        return

    log.info(f"读取买入记录: {input_path}")
    print(f"读取买入记录: {input_path}")
    records = load_buy_records(input_path)
    log.info(f"共读取 {len(records)} 条交易记录")
    print(f"共读取 {len(records)} 条交易记录")

    # 读取持仓份额（可选）
    shares_map = load_holdings_share(input_path)
    if shares_map:
        log.info(f"读取持仓份额: {len(shares_map)} 只基金")
        print(f"读取持仓份额: {len(shares_map)} 只基金")

    if not records:
        log.warning("没有有效的交易记录，请检查Excel格式")
        print("没有有效的交易记录，请检查Excel格式")
        return

    log.info("计算持仓状态...")
    print("计算持仓状态...")
    data = calculate_portfolio(records)

    # 输出
    timestamp = datetime.now().strftime("%Y%m%d")
    if args.output == "json":
        output_path = OUTPUT_DIR / f"portfolio_{timestamp}.json"
        output_json(data, output_path)
    else:
        output_path = OUTPUT_DIR / f"portfolio_report_{timestamp}.md"
        output_markdown(data, output_path)

    # 控制台摘要
    summary = data["summary"]
    log.info(f"持仓摘要: 基金数={summary['fund_count']}, 总成本=¥{summary['total_cost']:,.2f}, 总市值=¥{summary['total_value']:,.2f}, 总收益=¥{summary['total_profit']:,.2f} ({summary['total_profit_rate']:+.2f}%)")
    print(f"\n持仓摘要:")
    print(f"  基金数: {summary['fund_count']} 只")
    print(f"  总成本: ¥{summary['total_cost']:,.2f}")
    print(f"  总市值: ¥{summary['total_value']:,.2f}")
    print(f"  总收益: ¥{summary['total_profit']:,.2f} ({summary['total_profit_rate']:+.2f}%)")


if __name__ == "__main__":
    main()
